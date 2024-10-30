import pandas as pd
import os
from datetime import datetime
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Definir caminhos
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
PUBMED_FILE = os.path.join(BASE_DIR, "02_searches", "data", "pubmed_results_20240318.csv")
OUTPUT_FILE = os.path.join(BASE_DIR, "02_searches", "data", "screening_database.xlsx")

def process_references():
    # Ler arquivo original
    df = pd.read_csv(PUBMED_FILE)
    
    # Selecionar e renomear colunas relevantes
    columns_mapping = {
        'PMID': 'PMID',
        'Title': 'Título',
        'Authors': 'Autores',
        'Citation': 'Referência Completa',
        'First Author': 'Primeiro Autor',
    }
    
    # Opções para validação
    decisao_options = ['Incluir', 'Excluir', 'Dúvida']
    motivo_options = [
        'Não é ECR',
        'População inadequada',
        'Intervenção inadequada',
        'Desfechos inadequados',
        'Outro'
    ]
    
    # Criar Excel writer
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        # Aba de screening
        df_screening = df.rename(columns=columns_mapping)
        df_screening['Decisão'] = ''
        df_screening['Motivo_Exclusão'] = ''
        df_screening['Observações'] = ''
            
        df_screening.to_excel(writer, sheet_name='Screening', index=False)
        
        # Formatar planilha
        workbook = writer.book
        worksheet = writer.sheets['Screening']
        
        # Adicionar validação de dados
        decisao_validation = DataValidation(
            type="list",
            formula1=f'"{",".join(decisao_options)}"',
            allow_blank=True
        )
        
        motivo_validation = DataValidation(
            type="list",
            formula1=f'"{",".join(motivo_options)}"',
            allow_blank=True
        )
        
        # Aplicar validação às colunas
        decisao_col = get_column_letter(df_screening.columns.get_loc('Decisão') + 1)
        motivo_col = get_column_letter(df_screening.columns.get_loc('Motivo_Exclusão') + 1)
        
        worksheet.add_data_validation(decisao_validation)
        worksheet.add_data_validation(motivo_validation)
        
        decisao_validation.add(f'{decisao_col}2:{decisao_col}{len(df_screening) + 1}')
        motivo_validation.add(f'{motivo_col}2:{motivo_col}{len(df_screening) + 1}')
        
        # Ajustar largura das colunas
        for idx, col in enumerate(df_screening.columns):
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = 20
            
        # Formatar cabeçalho
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Aba de informações
        info_data = {
            'Data da Busca': [datetime.now().strftime('%Y-%m-%d')],
            'Base de Dados': ['PubMed'],
            'Total de Resultados': [2536],
            'Após Filtros': [58],
            'Filtros Aplicados': ['Clinical Trial, RCT, English/Portuguese/Spanish']
        }
        pd.DataFrame(info_data).to_excel(writer, sheet_name='Informações', index=False)
    
    return len(df)

if __name__ == "__main__":
    print("Processando referências...")
    n_refs = process_references()
    print(f"\nProcessamento concluído!")
    print(f"Total de referências processadas: {n_refs}")
    print(f"Arquivo salvo em: {OUTPUT_FILE}")
